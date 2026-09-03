import io
import os
import textwrap
from collections import defaultdict
from datetime import datetime

import pandas as pd
import openpyxl
import barcode as barcode_lib
from barcode.writer import ImageWriter
from PIL import Image, ImageDraw, ImageFont
from flask import Flask, request, render_template, send_file, flash, redirect, url_for
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "equivalencia-app-secret")

DATA_DIR = os.environ.get("DATA_DIR", "/app/data")
os.makedirs(DATA_DIR, exist_ok=True)

GRANDE_CACHE_PATH_XLS = os.path.join(DATA_DIR, "planilla_grande.xls")
GRANDE_CACHE_PATH_XLSX = os.path.join(DATA_DIR, "planilla_grande.xlsx")
GRANDE_CACHE_META = os.path.join(DATA_DIR, "planilla_grande.meta")

ALLOWED_EXT = {".xls", ".xlsx"}

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FONT_BOLD = os.path.join(BASE_DIR, "fonts", "DejaVuSans-Bold.ttf")
FONT_REGULAR = os.path.join(BASE_DIR, "fonts", "DejaVuSans.ttf")


def _ext(filename):
    return os.path.splitext(filename)[1].lower()


def _read_grande(path):
    """Lee la planilla grande (Equivalencia) sea .xls viejo (Crystal Reports) o .xlsx.
    Devuelve un dict: 'SKU COLOR' -> lista de (codigo_barra, talle, descripcion)
    """
    ext = _ext(path)
    engine = "xlrd" if ext == ".xls" else "openpyxl"
    xl = pd.ExcelFile(path, engine=engine)

    lookup = defaultdict(list)
    for sheet in xl.sheet_names:
        df = xl.parse(sheet, header=None, dtype=str)
        if df.empty:
            continue
        # La primera hoja trae una fila de encabezado ("Código", "Artículo", ...)
        if str(df.iloc[0, 0]).strip() == "Código":
            df = df.iloc[1:]
        for _, row in df.iterrows():
            codigo = row.get(0)
            articulo = row.get(1)
            descripcion = row.get(2)
            talle = row.get(5) if len(row) > 5 else None
            if pd.isna(codigo) or pd.isna(articulo):
                continue
            codigo = str(codigo).strip()
            articulo = str(articulo).strip()
            descripcion = "" if pd.isna(descripcion) else str(descripcion).strip()
            talle = "" if pd.isna(talle) else str(talle).strip()
            if not codigo or not articulo:
                continue
            lookup[articulo].append((codigo, talle, descripcion))
    return lookup


_LOOKUP_CACHE = {"path": None, "mtime": None, "lookup": None}


def _get_lookup(path):
    """Cachea en memoria el resultado de _read_grande mientras no cambie el archivo,
    para no re-parsear ~300k filas en cada pedido o cada etiqueta."""
    global _LOOKUP_CACHE
    mtime = os.path.getmtime(path)
    if _LOOKUP_CACHE["path"] == path and _LOOKUP_CACHE["mtime"] == mtime:
        return _LOOKUP_CACHE["lookup"]
    lookup = _read_grande(path)
    _LOOKUP_CACHE = {"path": path, "mtime": mtime, "lookup": lookup}
    return lookup


def _resolver_grande_cache():
    """Devuelve el path de la planilla grande cacheada en disco, o None si no hay ninguna."""
    if os.path.exists(GRANDE_CACHE_PATH_XLS):
        return GRANDE_CACHE_PATH_XLS
    if os.path.exists(GRANDE_CACHE_PATH_XLSX):
        return GRANDE_CACHE_PATH_XLSX
    return None


def _dedupe_codigos(codigos):
    """A veces la planilla grande tiene el mismo código de barra duplicado con
    distinto padding de ceros a la izquierda (ej. '00191743848643' y '191743848643').
    Los trata como el mismo código y se queda con la representación más corta."""
    vistos = {}
    for c in codigos:
        try:
            clave = int(c)
        except (TypeError, ValueError):
            clave = c
        if clave not in vistos or len(c) < len(vistos[clave]):
            vistos[clave] = c
    return list(vistos.values())


def _buscar_codigo(lookup, sku, color, talle):
    """Busca el código de barra exacto para SKU + Color + Talle."""
    if sku is None or color is None or str(sku).strip() == "" or str(color).strip() == "":
        return ""
    sku_str = str(sku).strip()
    if sku_str.endswith(".0"):
        sku_str = sku_str[:-2]
    color_str = str(color).strip()
    talle_str = "" if talle is None else str(talle).strip()

    key = f"{sku_str} {color_str}"
    candidatos = lookup.get(key, [])
    coincidencias = [
        codigo for codigo, t, _desc in candidatos if t.strip().upper() == talle_str.upper()
    ]
    coincidencias = _dedupe_codigos(coincidencias)

    if not coincidencias:
        return "SIN COINCIDENCIA"
    if len(coincidencias) == 1:
        return coincidencias[0]
    # Conflicto real de datos en la planilla grande: mismo SKU+Color+Talle con más
    # de un código de barra distinto. Se muestran todos para que se revise a mano.
    return " / ".join(coincidencias)


def _buscar_producto(lookup, sku, color, talle):
    """Busca código de barra + descripción exactos para SKU + Color + Talle.
    Devuelve (codigo, descripcion) o (None, None) si no hay coincidencia."""
    if not sku or not color:
        return None, None
    sku_str = str(sku).strip()
    if sku_str.endswith(".0"):
        sku_str = sku_str[:-2]
    color_str = str(color).strip()
    talle_str = "" if talle is None else str(talle).strip()

    key = f"{sku_str} {color_str}"
    candidatos = lookup.get(key, [])
    coincidencias = [
        (codigo, desc) for codigo, t, desc in candidatos if t.strip().upper() == talle_str.upper()
    ]
    if not coincidencias:
        return None, None

    codigos_dedup = _dedupe_codigos([c for c, _d in coincidencias])
    descripcion = coincidencias[0][1]
    # Para la etiqueta necesitamos un único código escaneable: si hay un conflicto real
    # de datos (mismo SKU+Color+Talle con más de un código), usamos el primero.
    codigo = codigos_dedup[0]
    return codigo, descripcion


def _buscar_columna(headers, *nombres_posibles):
    for nombre in nombres_posibles:
        if nombre in headers:
            return headers[nombre]
    return None


def _procesar(grande_path, pedido_path):
    """Abre el excel de pedido tal cual fue subido (mismo formato, mismo estilo) y
    completa la columna EQUIVALENCIA que ya viene en la planilla, buscando el
    código de barra exacto por SKU + Color + Talle. No inserta columnas nuevas ni
    toca el resto de la planilla."""
    lookup = _get_lookup(grande_path)

    wb = openpyxl.load_workbook(pedido_path)
    ws = wb.worksheets[0]

    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            headers[str(v).strip().upper()] = c

    col_equiv = _buscar_columna(headers, "EQUIVALENCIA")
    col_sku = _buscar_columna(headers, "SKU")
    col_color = _buscar_columna(headers, "COLOR")
    col_talle = _buscar_columna(headers, "TALLE")

    faltantes = [
        nombre
        for nombre, col in [
            ("EQUIVALENCIA", col_equiv),
            ("SKU", col_sku),
            ("COLOR", col_color),
            ("TALLE", col_talle),
        ]
        if col is None
    ]
    if faltantes:
        raise ValueError(
            "No encontré la columna "
            + ", ".join(faltantes)
            + " en la fila 1 de la planilla. Los encabezados deben llamarse exactamente así."
        )

    for row in range(2, ws.max_row + 1):
        sku = ws.cell(row=row, column=col_sku).value
        color = ws.cell(row=row, column=col_color).value
        talle = ws.cell(row=row, column=col_talle).value
        if sku is None and color is None:
            continue
        resultado = _buscar_codigo(lookup, sku, color, talle)
        cell = ws.cell(row=row, column=col_equiv, value=resultado)
        cell.number_format = "@"  # texto plano, para no perder ceros a la izquierda

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _generar_imagen_codigo_barra(codigo, module_height_mm=9.0, module_width_mm=0.28):
    """Genera la imagen del código de barra (EAN13 / UPC-A / Code128 según el largo)
    sin el texto legible debajo (eso lo dibujamos aparte con nuestra propia fuente)."""
    codigo = str(codigo).strip()
    solo_digitos = codigo.isdigit()
    if solo_digitos and len(codigo) == 13:
        clase = "ean13"
    elif solo_digitos and len(codigo) == 12:
        clase = "upca"
    elif solo_digitos and len(codigo) == 8:
        clase = "ean8"
    else:
        clase = "code128"

    BarcodeClass = barcode_lib.get_barcode_class(clase)
    bc = BarcodeClass(codigo, writer=ImageWriter())
    buf = io.BytesIO()
    bc.write(
        buf,
        options={
            "write_text": False,
            "quiet_zone": 1.0,
            "module_height": module_height_mm,
            "module_width": module_width_mm,
        },
    )
    buf.seek(0)
    return Image.open(buf).convert("L")


def _texto_ajustado(draw, texto, font, max_width_px, max_lineas=2):
    """Envuelve el texto para que entre en max_width_px, cortando con '…' si no entra."""
    palabras = texto.split()
    lineas = []
    actual = ""
    for palabra in palabras:
        prueba = (actual + " " + palabra).strip()
        ancho = draw.textbbox((0, 0), prueba, font=font)[2]
        if ancho <= max_width_px or not actual:
            actual = prueba
        else:
            lineas.append(actual)
            actual = palabra
            if len(lineas) == max_lineas - 1:
                break
    if actual:
        lineas.append(actual)
    lineas = lineas[:max_lineas]

    # Si sobró texto sin usar, agregamos "…" a la última línea que entre
    usado = " ".join(lineas)
    if len(usado) < len(texto):
        ultima = lineas[-1]
        while draw.textbbox((0, 0), ultima + "…", font=font)[2] > max_width_px and len(ultima) > 1:
            ultima = ultima[:-1]
        lineas[-1] = ultima.rstrip() + "…"
    return lineas


def generar_etiqueta(descripcion, color, talle, codigo, dpi=300):
    """Genera la etiqueta de 4cm x 2cm como imagen PNG (devuelve BytesIO)."""
    px_mm = dpi / 25.4
    ancho_mm, alto_mm = 40.0, 20.0
    W = round(ancho_mm * px_mm)
    H = round(alto_mm * px_mm)

    img = Image.new("L", (W, H), color=255)
    draw = ImageDraw.Draw(img)

    margen = round(1.3 * px_mm)
    col_izq_ancho = round(21.5 * px_mm)

    font_desc = ImageFont.truetype(FONT_REGULAR, size=round(2.3 * px_mm))
    font_color = ImageFont.truetype(FONT_REGULAR, size=round(2.6 * px_mm))
    font_talle = ImageFont.truetype(FONT_BOLD, size=round(4.6 * px_mm))
    font_digitos = ImageFont.truetype(FONT_REGULAR, size=round(2.0 * px_mm))

    y = margen
    max_w = col_izq_ancho - margen

    for linea in _texto_ajustado(draw, str(descripcion).upper(), font_desc, max_w, max_lineas=2):
        draw.text((margen, y), linea, font=font_desc, fill=0)
        y += draw.textbbox((0, 0), linea, font=font_desc)[3] + round(0.6 * px_mm)

    y += round(0.8 * px_mm)
    draw.text((margen, y), str(color).upper(), font=font_color, fill=0)
    y += draw.textbbox((0, 0), str(color).upper(), font=font_color)[3] + round(1.0 * px_mm)

    draw.text((margen, y), str(talle).upper(), font=font_talle, fill=0)

    # --- Barcode a la derecha ---
    codigo_str = str(codigo).strip()
    col_der_x = margen + col_izq_ancho + round(1.0 * px_mm)
    col_der_ancho = W - col_der_x - margen

    bc_img = _generar_imagen_codigo_barra(codigo_str)
    escala = col_der_ancho / bc_img.width
    bc_alto = min(round(bc_img.height * escala), round(13 * px_mm))
    bc_img = bc_img.resize((col_der_ancho, bc_alto))

    bc_y = margen
    img.paste(bc_img, (col_der_x, bc_y))

    digitos_y = bc_y + bc_alto + round(0.5 * px_mm)
    bbox = draw.textbbox((0, 0), codigo_str, font=font_digitos)
    digitos_x = col_der_x + max(0, (col_der_ancho - (bbox[2] - bbox[0])) // 2)
    draw.text((digitos_x, digitos_y), codigo_str, font=font_digitos, fill=0)

    salida = io.BytesIO()
    img.save(salida, format="PNG", dpi=(dpi, dpi))
    salida.seek(0)
    return salida


@app.route("/etiqueta", methods=["GET"])
def etiqueta():
    grande_cached = _resolver_grande_cache() is not None
    return render_template("etiqueta.html", grande_cached=grande_cached)


@app.route("/etiqueta/generar", methods=["GET", "POST"])
def etiqueta_generar():
    datos = request.values
    sku = (datos.get("sku") or "").strip()
    color = (datos.get("color") or "").strip()
    talle = (datos.get("talle") or "").strip()

    if not sku or not color or not talle:
        flash("Completá SKU, Color y Talle.")
        return redirect(url_for("etiqueta"))

    grande_path = _resolver_grande_cache()
    if not grande_path:
        flash("No hay planilla grande cargada todavía. Subila primero desde la página principal.")
        return redirect(url_for("etiqueta"))

    try:
        lookup = _get_lookup(grande_path)
        codigo, descripcion = _buscar_producto(lookup, sku, color, talle)
    except Exception as e:
        flash(f"Error buscando el producto: {e}")
        return redirect(url_for("etiqueta"))

    if not codigo:
        flash(f"No encontré ningún producto con SKU {sku}, Color {color} y Talle {talle} en la planilla grande.")
        return redirect(url_for("etiqueta"))

    try:
        imagen = generar_etiqueta(descripcion, color, talle, codigo)
    except Exception as e:
        flash(f"No pude generar la etiqueta: {e}")
        return redirect(url_for("etiqueta"))

    descargar = datos.get("descargar") == "1"
    nombre = f"etiqueta_{secure_filename(codigo)}.png"
    return send_file(
        imagen,
        mimetype="image/png",
        as_attachment=descargar,
        download_name=nombre,
    )


@app.route("/", methods=["GET"])
def index():
    grande_cached = os.path.exists(GRANDE_CACHE_META)
    grande_info = None
    if grande_cached:
        with open(GRANDE_CACHE_META) as f:
            grande_info = f.read().strip()
    return render_template("index.html", grande_cached=grande_cached, grande_info=grande_info)


@app.route("/procesar", methods=["POST"])
def procesar():
    pedido_file = request.files.get("pedido")
    grande_file = request.files.get("grande")

    if not pedido_file or pedido_file.filename == "":
        flash("Subí la planilla de pedido (la que hay que resolver).")
        return redirect(url_for("index"))

    if _ext(pedido_file.filename) != ".xlsx":
        flash("La planilla de pedido debe ser .xlsx (para poder mantener el mismo formato/estilo al insertar la columna).")
        return redirect(url_for("index"))

    # Resolver planilla grande: la subida ahora, o la que quedó cacheada
    if grande_file and grande_file.filename != "":
        if _ext(grande_file.filename) not in ALLOWED_EXT:
            flash("La planilla grande debe ser .xls o .xlsx")
            return redirect(url_for("index"))
        grande_ext = _ext(grande_file.filename)
        grande_path = GRANDE_CACHE_PATH_XLS if grande_ext == ".xls" else GRANDE_CACHE_PATH_XLSX
        # Limpiar el otro formato cacheado para no usar una versión vieja por error
        other_path = GRANDE_CACHE_PATH_XLSX if grande_ext == ".xls" else GRANDE_CACHE_PATH_XLS
        if os.path.exists(other_path):
            os.remove(other_path)
        grande_file.save(grande_path)
        with open(GRANDE_CACHE_META, "w") as f:
            f.write(f"{secure_filename(grande_file.filename)} — cargada {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    else:
        if os.path.exists(GRANDE_CACHE_PATH_XLS):
            grande_path = GRANDE_CACHE_PATH_XLS
        elif os.path.exists(GRANDE_CACHE_PATH_XLSX):
            grande_path = GRANDE_CACHE_PATH_XLSX
        else:
            flash("No hay planilla grande cargada todavía. Subí la planilla de Equivalencia primero.")
            return redirect(url_for("index"))

    pedido_bytes = io.BytesIO(pedido_file.read())
    pedido_bytes.name = pedido_file.filename

    try:
        output = _procesar(grande_path, pedido_bytes)
    except Exception as e:
        flash(f"Error procesando los archivos: {e}")
        return redirect(url_for("index"))

    nombre_salida = f"{secure_filename(os.path.splitext(pedido_file.filename)[0])} - con equivalencia.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=nombre_salida,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/health")
def health():
    return {"status": "ok"}


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 8000)))

